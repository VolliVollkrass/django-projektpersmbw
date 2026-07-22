from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse

# Templates nutzen {% static %}; im Testlauf gibt es kein WhiteNoise-Manifest
OHNE_MANIFEST = {
    "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
    "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
}

from personal.models import Mitarbeiter

from . import services
from .models import (
    Beihilfesatz,
    Besoldungsbetrag,
    Besoldungstabelle,
    Familienzuschlagtabelle,
    Familienzuschlagzeile,
    FZKinderErhoehung,
    Innenauftrag,
    Mietenstufe,
)


class InnenauftragKategorieTest(TestCase):
    def test_kategorie_aus_nummernpraefix(self):
        self.assertEqual(Innenauftrag(nummer="F031233-0034").kategorie, "F")
        self.assertEqual(Innenauftrag(nummer="A031210-0001").kategorie, "A")
        self.assertEqual(Innenauftrag(nummer="U031211-0001").kategorie, "U")
        self.assertEqual(Innenauftrag(nummer="X999").kategorie, "")


class BesoldungstabelleTest(TestCase):
    def setUp(self):
        self.alt = Besoldungstabelle.objects.create(gueltig_ab=date(2025, 4, 1))
        Besoldungsbetrag.objects.create(
            tabelle=self.alt, gruppe="A10", stufe=7, betrag=Decimal("4257.41")
        )

    def test_gueltig_fuer_stichtag(self):
        neu = Besoldungstabelle.objects.create(gueltig_ab=date(2026, 1, 1))
        self.assertEqual(Besoldungstabelle.gueltig_fuer(date(2025, 12, 31)), self.alt)
        self.assertEqual(Besoldungstabelle.gueltig_fuer(date(2026, 1, 1)), neu)
        self.assertIsNone(Besoldungstabelle.gueltig_fuer(date(2020, 1, 1)))

    def test_fortschreibung_rundet_kaufmaennisch(self):
        neu = services.besoldungstabelle_fortschreiben(date(2026, 1, 1), Decimal("2.6"))
        betrag = neu.betraege.get(gruppe="A10", stufe=7).betrag
        # 4257.41 * 1.026 = 4368.10 (wie ROUND(...;2) in der Excel-Vorlage)
        self.assertEqual(betrag, Decimal("4368.10"))
        self.assertEqual(neu.erhoehung_prozent, Decimal("2.6"))

    def test_fortschreibung_verlangt_juengeres_datum(self):
        with self.assertRaises(ValueError):
            services.besoldungstabelle_fortschreiben(date(2025, 4, 1), Decimal("2.6"))


class FZTabelleTest(TestCase):
    def test_fortschreibung_uebernimmt_leere_werte(self):
        alt = Familienzuschlagtabelle.objects.create(gueltig_ab=date(2025, 4, 1))
        Familienzuschlagzeile.objects.create(
            tabelle=alt, ortsklasse="V", stufe_v=Decimal("109.41"), stufe_l=None
        )
        FZKinderErhoehung.objects.create(
            tabelle=alt, ortsklasse="V", gruppe="A9", betrag=Decimal("100.00")
        )
        neu = services.fz_tabelle_fortschreiben(date(2026, 1, 1), Decimal("2.6"))
        zeile = neu.zeilen.get(ortsklasse="V")
        self.assertEqual(zeile.stufe_v, Decimal("112.25"))
        self.assertIsNone(zeile.stufe_l)
        self.assertEqual(
            neu.kinder_erhoehungen.get(ortsklasse="V", gruppe="A9").betrag,
            Decimal("102.60"),
        )


class BeihilfesatzTest(TestCase):
    def setUp(self):
        call_command("import_beihilfetarife", verbosity=0)

    def test_altersgruppen_lookup_tarif_830(self):
        self.assertEqual(Beihilfesatz.satz_fuer(2026, "830", 38), Decimal("183.74"))
        self.assertEqual(Beihilfesatz.satz_fuer(2026, "830", 39), Decimal("183.74"))
        self.assertEqual(Beihilfesatz.satz_fuer(2026, "830", 40), Decimal("210.93"))
        self.assertEqual(Beihilfesatz.satz_fuer(2026, "830", 85), Decimal("1050.61"))
        self.assertEqual(Beihilfesatz.kindersatz_fuer(2026, "830"), Decimal("182.64"))

    def test_pauschalbeitrag_tarif_814(self):
        self.assertEqual(Beihilfesatz.satz_fuer(2026, "814", 60), Decimal("2.15"))
        self.assertEqual(Beihilfesatz.satz_fuer(2026, "814", 70), Decimal("12.95"))
        self.assertIsNone(Beihilfesatz.kindersatz_fuer(2026, "814"))

    def test_unbekanntes_jahr_liefert_none(self):
        self.assertIsNone(Beihilfesatz.satz_fuer(1999, "830", 40))


@override_settings(STORAGES=OHNE_MANIFEST)
class StammdatenViewsTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user("test", password="x")
        self.client.force_login(self.user)

    def test_stammdaten_uebersicht(self):
        Besoldungstabelle.objects.create(gueltig_ab=date(2026, 1, 1))
        antwort = self.client.get(reverse("mbw_stammdaten"))
        self.assertEqual(antwort.status_code, 200)
        self.assertContains(antwort, "Besoldungstabellen")

    def test_mietenstufen_suche(self):
        Mietenstufe.objects.create(art="gemeinde", name="München", stufe="VII")
        antwort = self.client.get(reverse("mietenstufen"), {"q": "münchen"})
        self.assertContains(antwort, "VII")

    def test_debitor_anlegen_braucht_berechtigung(self):
        antwort = self.client.post(reverse("debitor_create"), {"name": "Test e. V."})
        self.assertEqual(antwort.status_code, 403)

    def test_besoldung_erhoehen_view(self):
        from django.contrib.auth.models import Permission

        self.user.user_permissions.add(
            Permission.objects.get(codename="add_besoldungstabelle")
        )
        Besoldungstabelle.objects.create(gueltig_ab=date(2025, 4, 1))
        self.client.post(
            reverse("besoldungstabelle_erhoehen"),
            {"besoldung-gueltig_ab": "2026-01-01", "besoldung-prozent": "2.6"},
        )
        self.assertTrue(
            Besoldungstabelle.objects.filter(gueltig_ab=date(2026, 1, 1)).exists()
        )


class MitarbeiterAbrechnungsfelderTest(TestCase):
    def test_neue_felder_speicherbar(self):
        mitarbeiter = Mitarbeiter.objects.create(
            personalnummer="99999",
            vorname="Test",
            nachname="Person",
            geburtsdatum=date(1980, 5, 1),
            ortsklasse="V",
            beihilfe_art=Mitarbeiter.BeihilfeArt.GESETZLICH,
            beihilfe_monat=Decimal("7.53"),
            kv_zuschuss_monat=Decimal("370.07"),
        )
        mitarbeiter.refresh_from_db()
        self.assertEqual(mitarbeiter.ortsklasse, "V")
        self.assertEqual(mitarbeiter.beihilfe_monat, Decimal("7.53"))


def _kalkulations_stammdaten():
    """Tabellenstände wie im Hochrechnungstool (Ausschnitt für A10/OK II)."""
    t1 = Besoldungstabelle.objects.create(gueltig_ab=date(2024, 11, 1))
    Besoldungsbetrag.objects.create(tabelle=t1, gruppe="A10", stufe=7, betrag=Decimal("4035.46"))
    t2 = Besoldungstabelle.objects.create(gueltig_ab=date(2025, 4, 1))
    Besoldungsbetrag.objects.create(tabelle=t2, gruppe="A10", stufe=7, betrag=Decimal("4257.41"))

    fz1 = Familienzuschlagtabelle.objects.create(gueltig_ab=date(2024, 11, 1))
    Familienzuschlagzeile.objects.create(
        tabelle=fz1, ortsklasse="II", stufe_v=Decimal("80.67"),
        stufe_1=Decimal("319.87"), stufe_2=Decimal("467.30"),
        kind_3=Decimal("456.92"), kind_weitere=Decimal("547.01"),
    )
    fz2 = Familienzuschlagtabelle.objects.create(gueltig_ab=date(2025, 4, 1))
    Familienzuschlagzeile.objects.create(
        tabelle=fz2, ortsklasse="II", stufe_v=Decimal("85.11"),
        stufe_1=Decimal("337.46"), stufe_2=Decimal("493.00"),
        kind_3=Decimal("482.05"), kind_weitere=Decimal("577.10"),
    )

    from .models import Zulagensatz
    Zulagensatz.objects.create(art="struktur", gueltig_ab=date(2024, 11, 1), betrag=Decimal("106.02"))
    Zulagensatz.objects.create(art="struktur", gueltig_ab=date(2025, 4, 1), betrag=Decimal("111.07"))

    Beihilfesatz.objects.create(jahr=2024, tarif="830", kategorie="erwachsen", alter_bis=39, satz=Decimal("162.71"))
    Beihilfesatz.objects.create(jahr=2024, tarif="830", kategorie="erwachsen", alter_bis=49, satz=Decimal("195.30"))
    Beihilfesatz.objects.create(jahr=2024, tarif="830", kategorie="kind", satz=Decimal("158.81"))


def _muster_einsatz(**mitarbeiter_felder):
    from decimal import Decimal as D
    from traeger.models import Einrichtung, Stelle, Traeger
    from einsatz.models import Einsatz

    traeger = Traeger.objects.create(name="Test-Träger", art="dw", traeger_id="T-1")
    einrichtung = Einrichtung.objects.create(
        traeger=traeger, name="Test-Einrichtung", einrichtungs_id="E-1",
        versorgungsumlage=D("0.40"),
    )
    stelle = Stelle.objects.create(einrichtung=einrichtung, name="Test-Stelle", stellen_id="S-1")
    felder = {
        "personalnummer": "88888", "vorname": "Eva", "nachname": "Mustermann",
        "geburtsdatum": date(1986, 6, 4), "familienstand": "verheiratet",
        "kinder": 2, "status_besoldung": "A10", "status_stufe": 7,
        "ortsklasse": "II", "beihilfe_art": Mitarbeiter.BeihilfeArt.PRIVAT_ALTER,
    }
    felder.update(mitarbeiter_felder)
    mitarbeiter = Mitarbeiter.objects.create(**felder)
    return Einsatz.objects.create(
        stelle=stelle, mitarbeiter=mitarbeiter,
        beginn=date(2025, 1, 1), umfang=D("1.00"), abrechnung=True,
    )


class HochrechnungTest(TestCase):
    """Verprobung gegen das Excel-Hochrechnungstool (Beispielfall Mustermann)."""

    def setUp(self):
        _kalkulations_stammdaten()
        self.einsatz = _muster_einsatz()

    def test_komponenten_wie_excel_beispiel(self):
        from .kalkulation import hochrechnung_berechnen

        summen = hochrechnung_berechnen(self.einsatz, 2025)["summen"]
        self.assertEqual(summen["grundbezug"], Decimal("50423.07"))
        self.assertEqual(summen["ofz"], Decimal("0.00"))
        self.assertEqual(summen["ofz_kinder"], Decimal("5838.90"))
        self.assertEqual(summen["strukturzulage"], Decimal("1317.69"))
        self.assertEqual(summen["sonderzahlung"], Decimal("3358.81"))
        self.assertEqual(summen["umlage"], Decimal("24375.39"))
        # Beihilfe: Excel-Tool rechnet nur den Erwachsenensatz (1952,52);
        # wir rechnen die Kinder mit (wie die echte PK-Erstattung): +2×158,81×12
        self.assertEqual(summen["beihilfe"], Decimal("5763.96"))
        self.assertEqual(summen["gesamt"], Decimal("91077.82"))

    def test_kinderlos_nutzt_stufe_v(self):
        from .kalkulation import hochrechnung_berechnen

        mitarbeiter = self.einsatz.mitarbeiter
        mitarbeiter.kinder = 0
        mitarbeiter.save()
        summen = hochrechnung_berechnen(self.einsatz, 2025)["summen"]
        # 3 × 80,67 + 9 × 85,11 = 1007,00
        self.assertEqual(summen["ofz"], Decimal("1008.00"))
        self.assertEqual(summen["ofz_kinder"], Decimal("0.00"))

    def test_fehlende_ortsklasse_meldet_fehler(self):
        from .kalkulation import KalkulationsFehler, hochrechnung_berechnen

        mitarbeiter = self.einsatz.mitarbeiter
        mitarbeiter.ortsklasse = ""
        mitarbeiter.save()
        with self.assertRaises(KalkulationsFehler) as kontext:
            hochrechnung_berechnen(self.einsatz, 2025)
        self.assertIn("Ortsklasse", kontext.exception.args[0][0])

    def test_gesetzlich_versichert_manuelle_betraege(self):
        from .kalkulation import hochrechnung_berechnen

        mitarbeiter = self.einsatz.mitarbeiter
        mitarbeiter.beihilfe_art = Mitarbeiter.BeihilfeArt.GESETZLICH
        mitarbeiter.beihilfe_monat = Decimal("7.53")
        mitarbeiter.beihilfe_kinder_monat = Decimal("7.28")
        mitarbeiter.kv_zuschuss_monat = Decimal("370.07")
        mitarbeiter.save()
        summen = hochrechnung_berechnen(self.einsatz, 2025)["summen"]
        self.assertEqual(summen["beihilfe"], Decimal("177.72"))     # (7,53+7,28)×12
        self.assertEqual(summen["kv_zuschuss"], Decimal("4440.84"))  # 370,07×12


class AbschlagVorschlagTest(TestCase):
    def test_abrundung_auf_500(self):
        from .kalkulation import abschlag_vorschlag

        self.assertEqual(abschlag_vorschlag(Decimal("87266.38")), Decimal("21500"))
        self.assertEqual(abschlag_vorschlag(Decimal("96000.00")), Decimal("24000"))
        self.assertEqual(abschlag_vorschlag(Decimal("1999.99")), Decimal("0"))


@override_settings(STORAGES=OHNE_MANIFEST)
class JahresakteViewTest(TestCase):
    def setUp(self):
        from django.contrib.auth.models import Permission

        _kalkulations_stammdaten()
        self.einsatz = _muster_einsatz()
        self.user = User.objects.create_user("sachbearbeitung", password="x")
        self.user.user_permissions.add(
            Permission.objects.get(codename="change_fakturierungsvorgang"),
            Permission.objects.get(codename="add_quartalsabrechnung"),
        )
        self.client.force_login(self.user)

    def test_kalkulieren_erzeugt_snapshot_und_abschlag(self):
        from .models import Fakturierungsvorgang

        self.client.post(reverse("jahresakte_kalkulieren", args=[self.einsatz.pk, 2025]))
        vorgang = Fakturierungsvorgang.objects.get(einsatz=self.einsatz, jahr=2025)
        self.assertEqual(vorgang.hochrechnung_gesamt, Decimal("91077.82"))
        self.assertEqual(vorgang.abschlag_quartal, Decimal("22500"))
        self.assertEqual(vorgang.kalkulation["summen"]["gesamt"], "91077.82")
        self.assertEqual(len(vorgang.kalkulation["monate"]), 12)

    def test_jahresakte_zeigt_rechnungstext(self):
        self.client.post(reverse("jahresakte_kalkulieren", args=[self.einsatz.pk, 2025]))
        antwort = self.client.get(reverse("jahresakte", args=[self.einsatz.pk, 2025]))
        self.assertEqual(antwort.status_code, 200)
        self.assertContains(antwort, "Pers.Nr.")
        self.assertContains(antwort, "wir bitten um unten aufgeführte Abschlagszahlung")
        self.assertContains(antwort, "Abschlagszahlung (1. von 4 Teilbeträgen)")
        self.assertContains(antwort, "22500,00 EUR")

    def test_quartal_buchen_mit_betrag_und_ruecksprung(self):
        from .models import Quartalsabrechnung

        antwort = self.client.post(
            reverse("quartalsabrechnung_buchen", args=[self.einsatz.pk]),
            {
                "jahr": "2025", "quartal": "1", "betrag": "22500",
                "next": reverse("jahresakte", args=[self.einsatz.pk, 2025]),
            },
        )
        self.assertEqual(antwort.url, reverse("jahresakte", args=[self.einsatz.pk, 2025]))
        eintrag = Quartalsabrechnung.objects.get(einsatz=self.einsatz, jahr=2025, quartal=1)
        self.assertEqual(eintrag.betrag, Decimal("22500"))


SAP_TESTDATEI = """01.11.2025          Dynamische Listenausgabe          1
\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t
\t\tPersNr\tNachname\tVorname\tL B E\tKostenst.\tAuftrag\tHauptb\tLArt\tLohnart-Langtext\t       Betrag\tWährg\tText\tFürper.\tInper.
\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t
\t\t88888\tMustermann\tEva\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t1\tGrundbezug\t4.805,20\tEUR\tHR 01/2025\t202501\t202501
\t\t88888\tMustermann\tEva\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t1\tGrundbezug\t-4.805,20\tEUR\tHR 02/2025\t202501\t202501
\t\t88888\tMustermann\tEva\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t1\tGrundbezug\t4.805,20\tEUR\tHR 02/2025\t202501\t202502
\t\t88888\tMustermann\tEva\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t3157\tFahrtkostenzuschuss stfr\t14,5\tEUR\tHR 01/2025\t202501\t202501
\t\t88888\tMustermann\tEva\tL_3-0120\t3-0312-033\tF031233-0034\t609800\t9015\tUmlage VF Brutto ab 2015\t1.922,08\tEUR\tHR 01/2025\t202501\t202501
\t\t99123\tFremd\tPerson\tL_3-0120\t3-0312-033\tX999999-0001\t603100\t1\tGrundbezug\t100,00\tEUR\tHR 01/2025\t202501\t202501
"""


# Rückrechnung mit Storno VOR Neubuchung + RV-Gegenbuchung über zwei Lohnarten.
SAP_STORNO_TESTDATEI = """01.11.2025          Dynamische Listenausgabe          1
\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t
\t\tPersNr\tNachname\tVorname\tL B E\tKostenst.\tAuftrag\tHauptb\tLArt\tLohnart-Langtext\t       Betrag\tWährg\tText\tFürper.\tInper.
\t\t\t\t\t\t\t\t\t\t\t\t\t\t\t
\t\t77777\tRück\tRita\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t0001\tGrundbezug\t-5.000,00\tEUR\tHR 08/2025\t202506\t202508
\t\t77777\tRück\tRita\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t0001\tGrundbezug\t5.000,00\tEUR\tHR 01/2026\t202506\t202601
\t\t77777\tRück\tRita\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t0001\tGrundbezug\t5.000,00\tEUR\tHR 07/2025\t202507\t202507
\t\t77777\tRück\tRita\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t1S32\tRentenvers.zuschlag\t700,00\tEUR\tHR 07/2025\t202507\t202507
\t\t77777\tRück\tRita\tL_3-0120\t3-0312-033\tF031233-0034\t603100\t9V98\tSV Beitrag AG Übernahme\t-700,00\tEUR\tHR 07/2025\t202507\t202507
"""


class SapImportTest(TestCase):
    def _import(self):
        from . import sap_import

        return sap_import.sap_import_anlegen(
            SAP_TESTDATEI.encode("utf-16"), "test.csv"
        )

    def test_storno_reihenfolgeunabhaengig_und_rv_gegenpaar(self):
        from . import sap_import
        from .models import PersonalkostenZeile

        lauf, statistik = sap_import.sap_import_anlegen(
            SAP_STORNO_TESTDATEI.encode("utf-16"), "storno.csv"
        )
        # 2 Paare: Rückrechnung (Storno vor Neubuchung) + RV-Gegenbuchung 1S32/9V98
        self.assertEqual(statistik["storno_paare"], 2)

        zeilen = list(PersonalkostenZeile.objects.filter(import_lauf=lauf).order_by("lfd_nr"))
        # Zeile 1 (−, früher) und Zeile 2 (+, später) heben sich trotz Reihenfolge auf
        self.assertEqual(zeilen[0].storno_partner, zeilen[1])
        # Zeile 3 (verbleibender Grundbezug) bleibt sichtbar
        self.assertIsNone(zeilen[2].storno_partner)
        # RV-Gegenbuchung: 1S32 (Zeile 4) <-> 9V98 (Zeile 5) über zwei Lohnarten
        self.assertEqual(zeilen[3].storno_partner, zeilen[4])

        fall = sap_import.faelle(lauf)[0]
        # Nur der eine verbleibende Grundbezug zählt; alle Nullpaare fallen raus.
        self.assertEqual(fall["bezuege"], Decimal("5000.00"))
        self.assertEqual(fall["gesamt"], Decimal("5000.00"))

    def test_parsen_storno_und_jahr(self):
        from .models import PersonalkostenZeile

        lauf, statistik = self._import()
        self.assertEqual(statistik["zeilen"], 6)
        self.assertEqual(statistik["storno_paare"], 1)
        self.assertEqual(lauf.jahr, 2025)

        zeilen = list(PersonalkostenZeile.objects.filter(import_lauf=lauf).order_by("lfd_nr"))
        self.assertEqual(zeilen[0].betrag, Decimal("4805.20"))
        # Zeile 1 (+) und Zeile 2 (−, gleiche Fürperiode) sind Partner
        self.assertEqual(zeilen[0].storno_partner, zeilen[1])
        self.assertEqual(zeilen[1].storno_partner, zeilen[0])
        # Zeile 3 (andere Für-/In-Periode-Kombination, kein Minus offen) bleibt frei
        self.assertIsNone(zeilen[2].storno_partner)

    def test_matching_und_fallsummen(self):
        from . import sap_import

        einsatz = _muster_einsatz()
        ia = Innenauftrag.objects.create(
            nummer="F031233-0034", einsatz=einsatz, andock_art="stelle"
        )
        lauf, _ = self._import()

        faelle = sap_import.faelle(lauf)
        self.assertEqual(len(faelle), 2)
        fall = next(f for f in faelle if f["personalnummer"] == "88888")
        self.assertEqual(fall["innenauftrag"], ia)
        self.assertEqual(fall["mitarbeiter"], einsatz.mitarbeiter)
        self.assertEqual(fall["bezuege"], Decimal("4819.70"))   # 4805,20 + 14,50 (Storno-Paar hebt sich auf)
        self.assertEqual(fall["umlage"], Decimal("1922.08"))
        fremd = next(f for f in faelle if f["personalnummer"] == "99123")
        self.assertIsNone(fremd["innenauftrag"])

    def test_kaputte_datei_wirft_fehler(self):
        from . import sap_import

        with self.assertRaises(sap_import.SapImportFehler):
            sap_import.sap_import_anlegen(b"nur quatsch ohne kopfzeile", "kaputt.txt")


@override_settings(STORAGES=OHNE_MANIFEST)
class SpitzabrechnungTest(TestCase):
    def setUp(self):
        from django.contrib.auth.models import Permission
        from . import sap_import

        _kalkulations_stammdaten()
        self.einsatz = _muster_einsatz()
        from .models import Debitor
        debitor = Debitor.objects.create(name="Träger e. V.", sap_nummer="1900000001")
        self.ia = Innenauftrag.objects.create(
            nummer="F031233-0034", einsatz=self.einsatz, andock_art="stelle", debitor=debitor
        )
        self.lauf, _ = sap_import.sap_import_anlegen(SAP_TESTDATEI.encode("utf-16"), "test.csv")

        self.user = User.objects.create_user("sachbearbeitung", password="x")
        self.user.user_permissions.add(
            Permission.objects.get(codename="change_fakturierungsvorgang"),
            Permission.objects.get(codename="add_quartalsabrechnung"),
        )
        self.client.force_login(self.user)

    def test_spitze_uebernehmen(self):
        from .models import Fakturierungsvorgang, Quartalsabrechnung

        # Kalkulieren (liefert Beihilfe für Position 10171) + ein Abschlag
        self.client.post(reverse("jahresakte_kalkulieren", args=[self.einsatz.pk, 2025]))
        Quartalsabrechnung.objects.create(
            einsatz=self.einsatz, jahr=2025, quartal=1, betrag=Decimal("1000.00")
        )

        self.client.post(reverse("jahresakte_spitze", args=[self.einsatz.pk, 2025]))
        vorgang = Fakturierungsvorgang.objects.get(einsatz=self.einsatz, jahr=2025)
        self.assertEqual(vorgang.pos_bezuege, Decimal("4819.70"))
        self.assertEqual(vorgang.pos_umlage, Decimal("1922.08"))
        self.assertEqual(vorgang.pos_beihilfe, Decimal("5763.96"))  # aus Kalkulation
        self.assertEqual(
            vorgang.spitze_rest,
            Decimal("4819.70") + Decimal("1922.08") + Decimal("5763.96") - Decimal("1000.00"),
        )
        self.assertEqual(vorgang.pk_import, self.lauf)

        antwort = self.client.get(reverse("jahresakte", args=[self.einsatz.pk, 2025]))
        self.assertContains(antwort, "Schlussrechnung 2025")
        self.assertContains(antwort, "10159")

    def test_fall_excel_download(self):
        import openpyxl
        from io import BytesIO

        antwort = self.client.get(
            reverse("pk_fall_excel", args=[self.lauf.pk, "88888", "F031233-0034"])
        )
        self.assertEqual(antwort.status_code, 200)
        self.assertIn("spreadsheetml", antwort["Content-Type"])
        wb = openpyxl.load_workbook(BytesIO(antwort.content))
        blatt = wb[wb.sheetnames[0]]
        self.assertIn("Mustermann", blatt["A1"].value)

    def test_debitor_mappe_download(self):
        import openpyxl
        from io import BytesIO

        debitor = self.ia.debitor
        antwort = self.client.get(
            reverse("pk_debitor_mappe", args=[self.lauf.pk, debitor.pk])
        )
        self.assertEqual(antwort.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(antwort.content))
        self.assertEqual(wb.sheetnames[0], "Übersicht")
        self.assertEqual(len(wb.sheetnames), 2)

    def test_import_detail_zeigt_unbekannten_auftrag(self):
        antwort = self.client.get(reverse("pk_import_detail", args=[self.lauf.pk]))
        self.assertEqual(antwort.status_code, 200)
        self.assertContains(antwort, "X999999-0001")
        self.assertContains(antwort, "unbekannt")


class DebitorNummerTest(TestCase):
    def test_extraktion_aus_zellwerten(self):
        from mbw.management.commands.import_innenauftragsliste import _debitor_nummer

        self.assertEqual(_debitor_nummer("1900032335"), "1900032335")
        # zwei Nummern: S4/HANA-Nummer (19…) gewinnt, egal in welcher Zeile
        self.assertEqual(_debitor_nummer("1600049958\n1900049958"), "1900049958")
        # zwei alte Nummern: die letzte gewinnt
        self.assertEqual(_debitor_nummer("1600036127\n1600028323"), "1600028323")
        # Nummer mit Vermerk
        self.assertEqual(_debitor_nummer("1600037504 - alt"), "1600037504")
        # reine Vermerke ergeben keine Nummer
        for vermerk in ("offen", "keinen", "DB LA", "DB Su-Ro", ""):
            self.assertEqual(_debitor_nummer(vermerk), "")


class ZahlungseingangStatusTest(TestCase):
    """Debitor-Status wird aus Forderung (Abschläge + Rest) und Zahlungen abgeleitet."""

    def setUp(self):
        from .models import Fakturierungsvorgang, Quartalsabrechnung

        self.einsatz = _muster_einsatz()
        self.vorgang = Fakturierungsvorgang.objects.create(einsatz=self.einsatz, jahr=2025)
        for quartal in (1, 2, 3):
            Quartalsabrechnung.objects.create(
                einsatz=self.einsatz, jahr=2025, quartal=quartal, betrag=Decimal("1000.00")
            )

    def test_status_ableitung(self):
        from .models import Fakturierungsvorgang, Zahlungseingang

        S = Fakturierungsvorgang.DebitorStatus
        self.assertEqual(self.vorgang.forderung_gesamt, Decimal("3000.00"))
        self.assertEqual(self.vorgang.status_ableiten(), S.OFFEN)

        Zahlungseingang.objects.create(vorgang=self.vorgang, betrag=Decimal("1000.00"))
        self.vorgang.debitor_status_aktualisieren()
        self.assertEqual(self.vorgang.status_ableiten(), S.TEILWEISE)
        self.assertEqual(self.vorgang.offener_betrag, Decimal("2000.00"))
        self.assertEqual(self.vorgang.debitor_status, S.TEILWEISE)

        Zahlungseingang.objects.create(vorgang=self.vorgang, betrag=Decimal("2000.00"))
        self.assertEqual(self.vorgang.status_ableiten(), S.AUSGEGLICHEN)
        self.assertEqual(self.vorgang.offener_betrag, Decimal("0.00"))

    def test_spitze_rest_erhoeht_forderung(self):
        self.vorgang.spitze_rest = Decimal("500.00")
        self.vorgang.save()
        self.assertEqual(self.vorgang.forderung_gesamt, Decimal("3500.00"))


@override_settings(STORAGES=OHNE_MANIFEST)
class ZahlungseingangViewTest(TestCase):
    def setUp(self):
        from django.contrib.auth.models import Permission
        from .models import Fakturierungsvorgang, Quartalsabrechnung

        self.einsatz = _muster_einsatz()
        Fakturierungsvorgang.objects.create(einsatz=self.einsatz, jahr=2025)
        Quartalsabrechnung.objects.create(
            einsatz=self.einsatz, jahr=2025, quartal=1, betrag=Decimal("1000.00")
        )
        self.user = User.objects.create_user("sb", password="x")
        self.user.user_permissions.add(
            Permission.objects.get(codename="add_zahlungseingang"),
            Permission.objects.get(codename="delete_zahlungseingang"),
        )
        self.client.force_login(self.user)

    def test_zahlung_anlegen_und_loeschen_aktualisiert_status(self):
        from .models import Fakturierungsvorgang, Zahlungseingang

        antwort = self.client.post(
            reverse("zahlungseingang_anlegen", args=[self.einsatz.pk, 2025]),
            {"datum": "2025-04-15", "betrag": "1000", "bemerkung": "Überweisung"},
        )
        self.assertRedirects(antwort, reverse("jahresakte", args=[self.einsatz.pk, 2025]))
        vorgang = Fakturierungsvorgang.objects.get(einsatz=self.einsatz, jahr=2025)
        self.assertEqual(vorgang.debitor_status, Fakturierungsvorgang.DebitorStatus.AUSGEGLICHEN)

        zahlung = Zahlungseingang.objects.get()
        self.assertEqual(zahlung.erfasst_von, self.user)
        self.client.post(reverse("zahlungseingang_loeschen", args=[zahlung.pk]))
        vorgang.refresh_from_db()
        self.assertEqual(Zahlungseingang.objects.count(), 0)
        self.assertEqual(vorgang.debitor_status, Fakturierungsvorgang.DebitorStatus.OFFEN)

    def test_jahresakte_zeigt_zahlungssektion(self):
        antwort = self.client.get(reverse("jahresakte", args=[self.einsatz.pk, 2025]))
        self.assertContains(antwort, "Zahlungseingänge")


@override_settings(STORAGES=OHNE_MANIFEST)
class CockpitTest(TestCase):
    def setUp(self):
        from .models import Fakturierungsvorgang, Quartalsabrechnung
        from django.utils.timezone import now

        self.einsatz = _muster_einsatz()  # beginn 2025-01-01, abrechnung=True
        vorgang = Fakturierungsvorgang.objects.create(
            einsatz=self.einsatz, jahr=2025, hochrechnung_gesamt=Decimal("80000.00"),
            kalkuliert_am=now(),
        )
        Quartalsabrechnung.objects.create(
            einsatz=self.einsatz, jahr=2025, quartal=1, betrag=Decimal("1500.00")
        )
        self.user = User.objects.create_user("leser", password="x")
        self.client.force_login(self.user)

    def test_cockpit_zeigt_fall_mit_pipeline(self):
        antwort = self.client.get(reverse("cockpit") + "?jahr=2025")
        self.assertEqual(antwort.status_code, 200)
        self.assertContains(antwort, "Mustermann, Eva")
        self.assertContains(antwort, "Jahres-Cockpit 2025")

    def test_einsatz_ausserhalb_jahr_nicht_gelistet(self):
        antwort = self.client.get(reverse("cockpit") + "?jahr=2024")
        self.assertNotContains(antwort, "Mustermann, Eva")

    def test_cockpit_export_liefert_xlsx(self):
        import io
        import openpyxl

        antwort = self.client.get(reverse("cockpit_export") + "?jahr=2025")
        self.assertEqual(antwort.status_code, 200)
        self.assertIn("spreadsheetml", antwort["Content-Type"])
        wb = openpyxl.load_workbook(io.BytesIO(antwort.content))
        ws = wb.active
        self.assertEqual(ws["A1"].value, "Jahres-Cockpit 2025 – Abrechnungsstatus")
        # Beispielperson taucht in Spalte A auf
        werte = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        self.assertIn("Mustermann, Eva", werte)


class StammdatenServiceTest(TestCase):
    def test_betrag_parsen(self):
        self.assertIsNone(services.betrag_parsen(""))
        self.assertIsNone(services.betrag_parsen(None))
        self.assertEqual(services.betrag_parsen("3.050,50"), Decimal("3050.50"))
        self.assertEqual(services.betrag_parsen("3050"), Decimal("3050.00"))
        with self.assertRaises(ValueError):
            services.betrag_parsen("abc")

    def test_besoldungstabelle_speichern_upsert_und_delete(self):
        tabelle = Besoldungstabelle.objects.create(gueltig_ab=date(2025, 1, 1))
        Besoldungsbetrag.objects.create(tabelle=tabelle, gruppe="A10", stufe=1, betrag=Decimal("3000"))
        Besoldungsbetrag.objects.create(tabelle=tabelle, gruppe="A10", stufe=2, betrag=Decimal("3100"))

        services.besoldungstabelle_speichern(
            tabelle, date(2025, 1, 1), "neu",
            {("A10", 1): Decimal("3050.00"), ("A10", 2): None, ("A11", 1): Decimal("3300.00")},
        )
        self.assertEqual(
            Besoldungsbetrag.objects.get(tabelle=tabelle, gruppe="A10", stufe=1).betrag,
            Decimal("3050.00"),
        )
        self.assertFalse(Besoldungsbetrag.objects.filter(tabelle=tabelle, gruppe="A10", stufe=2).exists())
        self.assertEqual(
            Besoldungsbetrag.objects.get(tabelle=tabelle, gruppe="A11", stufe=1).betrag,
            Decimal("3300.00"),
        )

    def test_gueltig_ab_kollision_wirft_fehler(self):
        Besoldungstabelle.objects.create(gueltig_ab=date(2024, 1, 1))
        tabelle = Besoldungstabelle.objects.create(gueltig_ab=date(2025, 1, 1))
        with self.assertRaises(ValueError):
            services.besoldungstabelle_speichern(tabelle, date(2024, 1, 1), "", {})


@override_settings(STORAGES=OHNE_MANIFEST)
class BesoldungEditorViewTest(TestCase):
    def setUp(self):
        from django.contrib.auth.models import Permission

        self.tabelle = Besoldungstabelle.objects.create(gueltig_ab=date(2025, 1, 1))
        Besoldungsbetrag.objects.create(tabelle=self.tabelle, gruppe="A10", stufe=1, betrag=Decimal("3000"))
        Besoldungsbetrag.objects.create(tabelle=self.tabelle, gruppe="A10", stufe=2, betrag=Decimal("3100"))
        self.user = User.objects.create_user("sb", password="x")
        self.user.user_permissions.add(
            Permission.objects.get(codename="change_besoldungstabelle"),
            Permission.objects.get(codename="add_besoldungstabelle"),
        )
        self.client.force_login(self.user)

    def test_direkt_bearbeiten_speichert(self):
        antwort = self.client.post(
            reverse("besoldungstabelle_detail", args=[self.tabelle.pk]),
            {"gueltig_ab": "2025-01-01", "bemerkung": "Korrektur", "stufen": "1,2",
             "b_A10_1": "3050,00", "b_A10_2": ""},
        )
        self.assertRedirects(antwort, reverse("besoldungstabelle_detail", args=[self.tabelle.pk]))
        self.assertEqual(
            Besoldungsbetrag.objects.get(tabelle=self.tabelle, gruppe="A10", stufe=1).betrag,
            Decimal("3050.00"),
        )
        self.assertFalse(Besoldungsbetrag.objects.filter(tabelle=self.tabelle, gruppe="A10", stufe=2).exists())

    def test_neue_tabelle_als_kopie(self):
        antwort = self.client.get(reverse("besoldungstabelle_neu"))
        self.assertEqual(antwort.status_code, 200)
        self.assertContains(antwort, "3000")  # aus jüngster Tabelle vorbelegt

        self.client.post(
            reverse("besoldungstabelle_neu"),
            {"gueltig_ab": "2026-01-01", "bemerkung": "", "stufen": "1,2",
             "b_A10_1": "3200,00", "b_A10_2": "3300,00"},
        )
        neu = Besoldungstabelle.objects.get(gueltig_ab=date(2026, 1, 1))
        self.assertEqual(
            Besoldungsbetrag.objects.get(tabelle=neu, gruppe="A10", stufe=1).betrag,
            Decimal("3200.00"),
        )

    def test_ohne_recht_kein_speichern(self):
        leser = User.objects.create_user("leser", password="x")
        self.client.force_login(leser)
        antwort = self.client.post(
            reverse("besoldungstabelle_detail", args=[self.tabelle.pk]),
            {"gueltig_ab": "2025-01-01", "stufen": "1,2", "b_A10_1": "9999"},
        )
        self.assertEqual(antwort.status_code, 403)
        self.assertEqual(
            Besoldungsbetrag.objects.get(tabelle=self.tabelle, gruppe="A10", stufe=1).betrag,
            Decimal("3000"),
        )


@override_settings(STORAGES=OHNE_MANIFEST)
class FzEditorViewTest(TestCase):
    def setUp(self):
        from django.contrib.auth.models import Permission

        self.tabelle = Familienzuschlagtabelle.objects.create(gueltig_ab=date(2025, 1, 1))
        Familienzuschlagzeile.objects.create(tabelle=self.tabelle, ortsklasse="II", stufe_l=Decimal("100"))
        self.user = User.objects.create_user("sb", password="x")
        self.user.user_permissions.add(
            Permission.objects.get(codename="change_familienzuschlagtabelle"),
        )
        self.client.force_login(self.user)

    def test_zeile_und_kindererhoehung_speichern(self):
        antwort = self.client.post(
            reverse("fz_tabelle_detail", args=[self.tabelle.pk]),
            {"gueltig_ab": "2025-01-01", "bemerkung": "", "z_II_stufe_l": "150", "k_A8_II": "50"},
        )
        self.assertRedirects(antwort, reverse("fz_tabelle_detail", args=[self.tabelle.pk]))
        zeile = Familienzuschlagzeile.objects.get(tabelle=self.tabelle, ortsklasse="II")
        self.assertEqual(zeile.stufe_l, Decimal("150.00"))
        self.assertEqual(
            FZKinderErhoehung.objects.get(tabelle=self.tabelle, gruppe="A8", ortsklasse="II").betrag,
            Decimal("50.00"),
        )
