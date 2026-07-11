"""Innenaufträge und Debitoren aus der Excel-Arbeitsliste importieren.

Erwartetes Blatt (Standard: erstes Blatt, z. B. „Tabelle1“ der Liste
„Innenaufträge-Diakone P3.2-2.xlsx“) mit Kopfzeile:
Kostenstelle | Auftrag | Name | ID | Erstattungspflichtiger Rechtsträger | Debitor … | Bemerkung

- Debitoren werden über die SAP-Nummer angelegt/aktualisiert (Name = Rechtsträger).
- Innenaufträge werden über die Auftragsnummer angelegt/aktualisiert
  (Kostenstelle, Debitor; „beendet“ im Namen -> aktiv = False).
- Bestehende Einsatz-Zuordnungen im CRM werden nicht angetastet.
"""

import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from mbw.models import Debitor, Innenauftrag

DEBITOR_NUMMER_MUSTER = re.compile(r"\d{7,20}")


def _debitor_nummer(wert):
    """SAP-Debitorennummer aus dem Zellinhalt ziehen.

    Die Spalte enthält teils Vermerke („offen“, „keinen“, „DB LA“) und teils
    alte und neue Nummer untereinander – bevorzugt wird die S4/HANA-Nummer
    (beginnt mit 19), sonst die letzte gefundene Nummer.
    """
    nummern = DEBITOR_NUMMER_MUSTER.findall(wert)
    if not nummern:
        return ""
    for nummer in reversed(nummern):
        if nummer.startswith("19"):
            return nummer
    return nummern[-1]

SPALTEN = {
    "kostenstelle": "kostenstelle",
    "auftrag": "auftrag",
    "name": "name",
    "erstattung": "rechtstraeger",
    "debitor": "debitor",
    "bemerkung": "bemerkung",
}


def _text(wert):
    if wert is None:
        return ""
    if isinstance(wert, float) and wert.is_integer():
        return str(int(wert))
    return str(wert).strip()


class Command(BaseCommand):
    help = "Importiert Innenaufträge + Debitoren aus der Arbeitsliste (xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("datei", help="Pfad zur Excel-Arbeitsliste")
        parser.add_argument("--blatt", default=None, help="Blattname (Standard: erstes Blatt)")
        parser.add_argument(
            "--dry-run", action="store_true", help="Nur prüfen, nichts speichern"
        )

    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["datei"], data_only=True)
        except Exception as exc:
            raise CommandError(f"Datei kann nicht gelesen werden: {exc}")
        ws = wb[options["blatt"]] if options["blatt"] else wb.worksheets[0]

        kopf = {}
        for zelle in next(ws.iter_rows(min_row=1, max_row=1)):
            text = " ".join(str(zelle.value or "").split()).lower()
            for schluessel, feld in SPALTEN.items():
                if text.startswith(schluessel):
                    kopf[feld] = zelle.column - 1
        fehlend = {"auftrag", "kostenstelle"} - set(kopf)
        if fehlend:
            raise CommandError(f"Kopfzeile unvollständig, fehlende Spalten: {', '.join(sorted(fehlend))}")

        try:
            with transaction.atomic():
                statistik = self._importieren(ws, kopf)
                if options["dry_run"]:
                    transaction.set_rollback(True)
        except Exception:
            raise

        modus = " (Probelauf, nichts gespeichert)" if options["dry_run"] else ""
        self.stdout.write(self.style.SUCCESS(
            f"{statistik['auftraege_neu']} Innenaufträge neu, {statistik['auftraege_aktualisiert']} aktualisiert, "
            f"{statistik['debitoren']} Debitoren angelegt/aktualisiert{modus}"
        ))
        if statistik["ohne_debitor"]:
            self.stdout.write(self.style.WARNING(
                f"{len(statistik['ohne_debitor'])} Aufträge mit Rechtsträger, aber ohne Debitorennummer "
                f"(Debitor bitte manuell zuordnen): {', '.join(statistik['ohne_debitor'][:15])}"
                + (" …" if len(statistik["ohne_debitor"]) > 15 else "")
            ))

    def _importieren(self, ws, kopf):
        statistik = {
            "auftraege_neu": 0,
            "auftraege_aktualisiert": 0,
            "debitoren": 0,
            "ohne_debitor": [],
        }
        debitoren_cache = {}

        def wert(zeile, feld):
            index = kopf.get(feld)
            return _text(zeile[index]) if index is not None and index < len(zeile) else ""

        for zeile in ws.iter_rows(min_row=2, values_only=True):
            nummer = wert(zeile, "auftrag")
            if not nummer:
                continue

            name = wert(zeile, "name")
            rechtstraeger = wert(zeile, "rechtstraeger")
            debitor_nr_roh = wert(zeile, "debitor")
            debitor_nr = _debitor_nummer(debitor_nr_roh)
            bemerkung = wert(zeile, "bemerkung")

            debitor = None
            if debitor_nr:
                debitor = debitoren_cache.get(debitor_nr)
                if debitor is None:
                    vorlage = {"name": (rechtstraeger or debitor_nr)[:200]}
                    if debitor_nr_roh != debitor_nr:
                        vorlage["bemerkung"] = f"Debitor-Spalte der Arbeitsliste: {debitor_nr_roh}"
                    debitor, _ = Debitor.objects.update_or_create(
                        sap_nummer=debitor_nr, defaults=vorlage
                    )
                    debitoren_cache[debitor_nr] = debitor
                    statistik["debitoren"] += 1
            elif rechtstraeger or debitor_nr_roh:
                statistik["ohne_debitor"].append(
                    f"{nummer}{f' („{debitor_nr_roh}“)' if debitor_nr_roh else ''}"
                )

            beendet = "beendet" in name.lower()
            felder = {"kostenstelle": wert(zeile, "kostenstelle")}
            if debitor:
                felder["debitor"] = debitor
            if beendet:
                felder["aktiv"] = False

            auftrag, neu = Innenauftrag.objects.get_or_create(
                nummer=nummer,
                defaults={
                    **felder,
                    "bezeichnung": name,
                    "bemerkung": bemerkung,
                },
            )
            if neu:
                statistik["auftraege_neu"] += 1
            else:
                for feld, inhalt in felder.items():
                    setattr(auftrag, feld, inhalt)
                if not auftrag.bezeichnung:
                    auftrag.bezeichnung = name
                if bemerkung and bemerkung not in auftrag.bemerkung:
                    auftrag.bemerkung = f"{auftrag.bemerkung}\n{bemerkung}".strip()
                auftrag.save()
                statistik["auftraege_aktualisiert"] += 1

        return statistik
