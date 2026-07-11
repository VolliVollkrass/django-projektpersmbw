"""Stammdaten aus dem Excel-Hochrechnungstool des LKA importieren.

Liest aus der Vorlage (xltx/xlsx, berechnete Werte):
- datierte Blätter (z. B. "01.01.26")  -> Besoldungstabellen + Zulagensätze
- "FZ "-Blätter                        -> Familienzuschlagtabellen
- Blatt "Beihilfe"                     -> Beihilfesätze Tarif 830 des Jahres
- Blatt "Ortsklasse"                   -> Mietenstufen (Gemeinden + Landkreise)

Mehrfach ausführbar (update_or_create über die fachlichen Schlüssel).
"""

import re
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from mbw.models import (
    Beihilfesatz,
    Besoldungsbetrag,
    Besoldungstabelle,
    Familienzuschlagtabelle,
    Familienzuschlagzeile,
    FZKinderErhoehung,
    Mietenstufe,
    Zulagensatz,
)

ORTSKLASSEN = {"I", "II", "III", "IV", "V", "VI", "VII"}
GRUPPEN_MUSTER = re.compile(r"^A\s?(\d{1,2})$")
AMTSZULAGE_MUSTER = re.compile(r"^(A\d{1,2})\s+([IVX]+)$")

FZ_SPALTEN = {
    "stufe l": "stufe_l",
    "stufe v": "stufe_v",
    "stufe 1": "stufe_1",
    "stufe 2": "stufe_2",
    "zzgl. für das 3. kind": "kind_3",
    "zzgl. je weiterem kind": "kind_weitere",
}


def _dezimal(wert):
    return Decimal(str(round(float(wert), 2)))


def _prozent(wert):
    """Erhöhung liegt in der Vorlage als Faktor (0.026) vor.

    Manche Blätter tragen stattdessen einen absoluten Euro-Betrag (z. B. 200)
    ein – alles über 100 % ist offensichtlich kein Prozentsatz und wird
    verworfen.
    """
    if wert is None:
        return None
    prozent = round(float(wert) * 100, 2)
    if not 0 <= prozent <= 100:
        return None
    return Decimal(str(prozent))


def _gueltig_ab(ws):
    for row in ws.iter_rows(max_row=3):
        for zelle in row:
            if isinstance(zelle.value, datetime):
                return zelle.value.date()
            if isinstance(zelle.value, date):
                return zelle.value
    # Fallback: Blattname ist das Datum (z. B. "01.12.2022", "FZ 01.04.2023")
    name = ws.title.strip().removeprefix("FZ ").strip()
    for format_ in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(name, format_).date()
        except ValueError:
            continue
    return None


def _erhoehung(ws):
    for row in ws.iter_rows(max_row=4):
        werte = [z.value for z in row]
        for i, wert in enumerate(werte):
            if isinstance(wert, str) and wert.strip().lower() == "erhöhung":
                for folge in werte[i + 1:]:
                    if isinstance(folge, (int, float)):
                        return _prozent(folge)
    return None


class Command(BaseCommand):
    help = "Importiert Besoldungs-, Familienzuschlag-, Beihilfe- und Ortsklassen-Stammdaten aus dem Hochrechnungstool (xltx)."

    def add_arguments(self, parser):
        parser.add_argument("datei", help="Pfad zum Hochrechnungstool (xltx/xlsx)")

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            wb = openpyxl.load_workbook(options["datei"], data_only=True)
        except Exception as exc:
            raise CommandError(f"Datei kann nicht gelesen werden: {exc}")

        for name in wb.sheetnames:
            if name.lower().startswith("fz "):
                self._fz_blatt(wb[name])
            elif re.match(r"^\d{2}\.\d{2}\.\d{2,4}$", name.strip()):
                self._besoldung_blatt(wb[name])

        if "Beihilfe" in wb.sheetnames:
            self._beihilfe_blatt(wb["Beihilfe"])
        if "Ortsklasse" in wb.sheetnames:
            self._ortsklasse_blatt(wb["Ortsklasse"])

        self.stdout.write(self.style.SUCCESS("Import abgeschlossen."))

    # --- Besoldung ---------------------------------------------------------

    def _besoldung_blatt(self, ws):
        gueltig_ab = _gueltig_ab(ws)
        if not gueltig_ab:
            self.stdout.write(self.style.WARNING(f"[{ws.title}] kein Gültig-ab-Datum gefunden – übersprungen"))
            return

        stufen_spalten, kopfzeile = self._stufen_kopf(ws)
        if not stufen_spalten:
            self.stdout.write(self.style.WARNING(f"[{ws.title}] keine Stufen-Kopfzeile gefunden – übersprungen"))
            return

        tabelle, _ = Besoldungstabelle.objects.update_or_create(
            gueltig_ab=gueltig_ab,
            defaults={"erhoehung_prozent": _erhoehung(ws), "bemerkung": f"Import aus Hochrechnungstool, Blatt „{ws.title}“"},
        )

        anzahl = 0
        for row in ws.iter_rows(min_row=kopfzeile + 1):
            treffer = GRUPPEN_MUSTER.match(str(row[0].value).strip()) if row[0].value else None
            if not treffer:
                continue
            gruppe = f"A{treffer.group(1)}"
            for zelle in row:
                stufe = stufen_spalten.get(zelle.column)
                if stufe and isinstance(zelle.value, (int, float)):
                    Besoldungsbetrag.objects.update_or_create(
                        tabelle=tabelle, gruppe=gruppe, stufe=stufe,
                        defaults={"betrag": _dezimal(zelle.value)},
                    )
                    anzahl += 1

        self._zulagen(ws, gueltig_ab)
        self.stdout.write(f"[{ws.title}] Besoldungstabelle ab {gueltig_ab:%d.%m.%Y}: {anzahl} Beträge")

    def _stufen_kopf(self, ws):
        """Zeile mit den Stufennummern 1–11 finden -> {Spalte: Stufe}."""
        for row in ws.iter_rows(max_row=15):
            ganze = [
                (z.column, int(z.value))
                for z in row
                if isinstance(z.value, (int, float)) and not isinstance(z.value, bool)
                and float(z.value).is_integer() and 1 <= z.value <= 11
            ]
            if len(ganze) >= 8 and [s for _, s in ganze] == sorted({s for _, s in ganze}):
                return dict(ganze), row[0].row
        return {}, None

    def _zulagen(self, ws, gueltig_ab):
        for row in ws.iter_rows():
            for zelle in row:
                wert = zelle.value
                # Strukturzulage: Label "A 9 bis A 13", Betrag rechts daneben
                if isinstance(wert, str) and "a 9 bis a 13" in wert.strip().lower():
                    betrag = self._naechste_zahl(ws, zelle)
                    if betrag is not None:
                        Zulagensatz.objects.update_or_create(
                            art=Zulagensatz.Art.STRUKTUR, variante="", gueltig_ab=gueltig_ab,
                            defaults={"betrag": betrag},
                        )
                # Amtszulagen: Zeilen "A9 I", "A10 I", ... Betrag rechts daneben
                if zelle.column == 1 and isinstance(wert, str):
                    treffer = AMTSZULAGE_MUSTER.match(wert.strip())
                    if treffer:
                        betrag = self._naechste_zahl(ws, zelle)
                        if betrag is not None:
                            Zulagensatz.objects.update_or_create(
                                art=Zulagensatz.Art.AMT,
                                variante=f"{treffer.group(1)} {treffer.group(2)}",
                                gueltig_ab=gueltig_ab,
                                defaults={"betrag": betrag},
                            )

    @staticmethod
    def _naechste_zahl(ws, zelle):
        for spalte in range(zelle.column + 1, zelle.column + 4):
            wert = ws.cell(row=zelle.row, column=spalte).value
            if isinstance(wert, (int, float)):
                return _dezimal(wert)
        return None

    # --- Familienzuschlag --------------------------------------------------

    def _fz_blatt(self, ws):
        gueltig_ab = _gueltig_ab(ws)
        if not gueltig_ab:
            self.stdout.write(self.style.WARNING(f"[{ws.title}] kein Gültig-ab-Datum gefunden – übersprungen"))
            return

        spalten, kopfzeile = self._fz_kopf(ws)
        if not spalten:
            self.stdout.write(self.style.WARNING(f"[{ws.title}] keine OFZ-Kopfzeile gefunden – übersprungen"))
            return

        tabelle, _ = Familienzuschlagtabelle.objects.update_or_create(
            gueltig_ab=gueltig_ab,
            defaults={"erhoehung_prozent": _erhoehung(ws), "bemerkung": f"Import aus Hochrechnungstool, Blatt „{ws.title}“"},
        )

        # In der Vorlage gelten Beträge über verbundene Zellen für mehrere
        # Ortsklassen (z. B. Stufe V ein Wert für I–IV) – Anker-Wert auffüllen.
        wert_matrix = {}
        for bereich in ws.merged_cells.ranges:
            anker = ws.cell(row=bereich.min_row, column=bereich.min_col).value
            if isinstance(anker, (int, float)):
                for zeile_nr in range(bereich.min_row, bereich.max_row + 1):
                    for spalte_nr in range(bereich.min_col, bereich.max_col + 1):
                        wert_matrix[(zeile_nr, spalte_nr)] = anker

        zeilen = 0
        for row in ws.iter_rows(min_row=kopfzeile + 1, max_row=kopfzeile + 10):
            ortsklasse = str(row[0].value).strip() if row[0].value else ""
            if ortsklasse not in ORTSKLASSEN:
                continue
            werte = {}
            for zelle in row:
                feld = spalten.get(zelle.column)
                if not feld:
                    continue
                wert = zelle.value
                if not isinstance(wert, (int, float)):
                    wert = wert_matrix.get((zelle.row, zelle.column))
                if isinstance(wert, (int, float)):
                    werte[feld] = _dezimal(wert)
                else:
                    werte[feld] = None
            Familienzuschlagzeile.objects.update_or_create(
                tabelle=tabelle, ortsklasse=ortsklasse, defaults=werte
            )
            zeilen += 1

        erhoehungen = self._fz_kinder_erhoehungen(ws, tabelle)
        self.stdout.write(
            f"[{ws.title}] OFZ-Tabelle ab {gueltig_ab:%d.%m.%Y}: {zeilen} Ortsklassen, {erhoehungen} Kindererhöhungen"
        )

    def _fz_kopf(self, ws):
        for row in ws.iter_rows(max_row=15):
            spalten = {}
            for zelle in row:
                if isinstance(zelle.value, str):
                    feld = FZ_SPALTEN.get(" ".join(zelle.value.split()).lower())
                    if feld:
                        spalten[zelle.column] = feld
            if len(spalten) >= 4:
                return spalten, row[0].row
        return {}, None

    def _fz_kinder_erhoehungen(self, ws, tabelle):
        """Block „Erhöhungsbeträge A 8 bis A 10“: Kopfzeile A 8/A 9/A 10, darunter Ortsklassen."""
        anzahl = 0
        for row in ws.iter_rows():
            gruppen = {
                z.column: f"A{GRUPPEN_MUSTER.match(str(z.value).strip()).group(1)}"
                for z in row
                if isinstance(z.value, str) and GRUPPEN_MUSTER.match(z.value.strip())
            }
            if len(gruppen) < 3:
                continue
            for folge in ws.iter_rows(min_row=row[0].row + 1, max_row=row[0].row + 10):
                ortsklasse = str(folge[0].value).strip() if folge[0].value else ""
                if ortsklasse not in ORTSKLASSEN:
                    continue
                for zelle in folge:
                    gruppe = gruppen.get(zelle.column)
                    if gruppe and isinstance(zelle.value, (int, float)):
                        FZKinderErhoehung.objects.update_or_create(
                            tabelle=tabelle, ortsklasse=ortsklasse, gruppe=gruppe,
                            defaults={"betrag": _dezimal(zelle.value)},
                        )
                        anzahl += 1
            break
        return anzahl

    # --- Beihilfe (Tarif 830) ----------------------------------------------

    def _beihilfe_blatt(self, ws):
        jahr = ws["A1"].value
        if not isinstance(jahr, int):
            self.stdout.write(self.style.WARNING("[Beihilfe] kein Jahr in A1 – übersprungen"))
            return

        zeilen = []
        for row in ws.iter_rows(min_row=2):
            beschriftung, satz = row[1].value if len(row) > 1 else None, row[2].value if len(row) > 2 else None
            if not isinstance(satz, (int, float)):
                continue
            if isinstance(beschriftung, str) and beschriftung.strip().lower() == "kinder":
                zeilen.append((Beihilfesatz.Kategorie.KIND, None, _dezimal(satz)))
            elif isinstance(beschriftung, (int, float)):
                zeilen.append((Beihilfesatz.Kategorie.ERWACHSEN, int(beschriftung), _dezimal(satz)))

        # höchste Altersangabe ist „ab X Jahren“ -> keine Obergrenze
        erwachsene = [z for z in zeilen if z[0] == Beihilfesatz.Kategorie.ERWACHSEN]
        max_alter = max((z[1] for z in erwachsene), default=None)
        anzahl = 0
        for kategorie, alter, satz in zeilen:
            alter_bis = None if (kategorie == Beihilfesatz.Kategorie.ERWACHSEN and alter == max_alter) else alter
            Beihilfesatz.objects.update_or_create(
                jahr=jahr, tarif=Beihilfesatz.Tarif.T830, kategorie=kategorie, alter_bis=alter_bis,
                defaults={"satz": satz},
            )
            anzahl += 1
        self.stdout.write(f"[Beihilfe] Tarif 830 für {jahr}: {anzahl} Sätze")

    # --- Ortsklassen / Mietenstufen ----------------------------------------

    def _ortsklasse_blatt(self, ws):
        gemeinden = kreise = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_g, stufe_g = (row[0], row[1]) if len(row) > 1 else (None, None)
            name_k, stufe_k = (row[3], row[4]) if len(row) > 4 else (None, None)
            if isinstance(name_g, str) and str(stufe_g).strip() in ORTSKLASSEN:
                Mietenstufe.objects.update_or_create(
                    art=Mietenstufe.Art.GEMEINDE, name=name_g.strip(),
                    defaults={"stufe": str(stufe_g).strip()},
                )
                gemeinden += 1
            if isinstance(name_k, str) and str(stufe_k).strip() in ORTSKLASSEN:
                Mietenstufe.objects.update_or_create(
                    art=Mietenstufe.Art.KREIS, name=name_k.strip(),
                    defaults={"stufe": str(stufe_k).strip()},
                )
                kreise += 1
        self.stdout.write(f"[Ortsklasse] {gemeinden} Gemeinden, {kreise} Landkreise")
