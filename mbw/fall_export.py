"""Fall-Excel im Format der bisherigen Hand-Auswertung („Gründel-Format“).

Je Fall (Person × Innenauftrag): alle SAP-Zeilen nach Lohnart gruppiert,
Storno-Paare markiert, Summenzeile je Lohnart, unten die Zusammenfassung mit
den drei SAP-Rechnungspositionen (10159/10167/10171), Abschlägen und Rest.
Zusätzlich eine Sammelmappe je Debitor (Deckblatt + ein Blatt pro Fall).
ELKB-CI-Farben wie in datenaustausch/xlsx.py.
"""

import io
import re
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from datenaustausch.xlsx import (
    ELKB_BLAUGRAU,
    ELKB_VIOLETT,
    ELKB_VIOLETT_HELL,
    ELKB_VIOLETT_MITTEL,
)

from . import sap_import
from .models import Fakturierungsvorgang, Quartalsabrechnung

KOPF_FUELLUNG = PatternFill("solid", fgColor=ELKB_VIOLETT)
KOPF_SCHRIFT = Font(bold=True, color="FFFFFF")
SUMME_FUELLUNG = PatternFill("solid", fgColor=ELKB_VIOLETT_MITTEL)
SUMME_SCHRIFT = Font(bold=True, color=ELKB_VIOLETT)
BLOCK_FUELLUNG = PatternFill("solid", fgColor=ELKB_VIOLETT_HELL)
STORNO_SCHRIFT = Font(italic=True, color="9CA3AF")
TITEL_SCHRIFT = Font(bold=True, size=14, color=ELKB_VIOLETT)
BETRAG_FORMAT = "#,##0.00"

SPALTEN = [
    ("PersNr", "personalnummer", 8),
    ("Nachname", "nachname", 14),
    ("Vorname", "vorname", 12),
    ("L B E", "lbe", 10),
    ("Kostenst.", "kostenstelle", 12),
    ("Auftrag", "auftrag_nummer", 14),
    ("Hauptb", "hauptbuchkonto", 8),
    ("LArt", "lohnart", 7),
    ("Lohnart-Langtext", "lohnart_text", 26),
    ("Betrag", "betrag", 12),
    ("Währg", "waehrung", 6),
    ("Text", "text", 11),
    ("Fürper.", "fuer_periode", 8),
    ("Inper.", "in_periode", 8),
    ("lfd.Nr.", "lfd_nr", 8),
    ("Prüfung", None, 11),
    ("Quelle", None, 16),
]
BETRAG_SPALTE = 10


def _blattname(fall, jahr):
    name = f"{fall['nachname']}_{fall['vorname']}_{fall['personalnummer']}"
    return re.sub(r"[\\/*?:\[\]]", "_", name)[:31]


def _zusammenfassung(fall, vorgang, abschlaege):
    """Positionsbeträge + Abschläge/Rest; Beihilfe kommt aus der Kalkulation."""
    beihilfe = kv = None
    if vorgang and vorgang.kalkulation:
        beihilfe = Decimal(vorgang.kalkulation["summen"]["beihilfe"])
        kv = Decimal(vorgang.kalkulation["summen"]["kv_zuschuss"])
    pos_beihilfe = (beihilfe or Decimal("0")) + (kv or Decimal("0"))
    pk_gesamt = fall["bezuege"] + fall["umlage"] + pos_beihilfe
    abschlag_summe = sum((a.betrag or Decimal("0") for a in abschlaege), Decimal("0"))
    return {
        "beihilfe": beihilfe,
        "kv": kv,
        "pos_beihilfe": pos_beihilfe,
        "pk_gesamt": pk_gesamt,
        "abschlag_summe": abschlag_summe,
        "abschlag_anzahl": len(abschlaege),
        "rest": pk_gesamt - abschlag_summe,
    }


def _betrag_zelle(ws, zeile, spalte, wert, fett=False):
    zelle = ws.cell(row=zeile, column=spalte, value=wert)
    zelle.number_format = BETRAG_FORMAT
    if fett:
        zelle.font = SUMME_SCHRIFT
    return zelle


def fall_blatt(wb, import_lauf, fall, vorgang=None, abschlaege=()):
    ws = wb.create_sheet(title=_blattname(fall, import_lauf.jahr))

    ws.cell(row=1, column=1, value=(
        f"PK-Abrechnung {fall['vorname']} {fall['nachname']} "
        f"#{fall['personalnummer']} {import_lauf.jahr}"
    )).font = TITEL_SCHRIFT

    for spalte, (beschriftung, _, breite) in enumerate(SPALTEN, start=1):
        zelle = ws.cell(row=3, column=spalte, value=beschriftung)
        zelle.font = KOPF_SCHRIFT
        zelle.fill = KOPF_FUELLUNG
        ws.column_dimensions[get_column_letter(spalte)].width = breite
    ws.freeze_panes = "A4"

    zeilen_nr = 4
    lfd_zu_zeile = {}
    gruppen = sap_import.fall_zeilen(import_lauf, fall["personalnummer"], fall["auftrag_nummer"])
    gesamt = Decimal("0")

    for gruppe in gruppen:
        for eintrag in gruppe["zeilen"]:
            lfd_zu_zeile[eintrag.lfd_nr] = zeilen_nr
            for spalte, (_, feld, _) in enumerate(SPALTEN, start=1):
                if feld is None:
                    continue
                wert = getattr(eintrag, feld)
                zelle = ws.cell(row=zeilen_nr, column=spalte, value=wert)
                if feld == "betrag":
                    zelle.number_format = BETRAG_FORMAT
            if eintrag.storno_partner_id:
                ws.cell(row=zeilen_nr, column=16, value="ausblenden")
                ws.cell(row=zeilen_nr, column=17, value=f"Partner: Nr.{eintrag.storno_partner.lfd_nr}")
                for spalte in range(1, len(SPALTEN) + 1):
                    ws.cell(row=zeilen_nr, column=spalte).font = STORNO_SCHRIFT
            zeilen_nr += 1

        ws.cell(row=zeilen_nr, column=8, value=gruppe["lohnart"]).font = SUMME_SCHRIFT
        ws.cell(row=zeilen_nr, column=9, value=gruppe["lohnart_text"]).font = SUMME_SCHRIFT
        _betrag_zelle(ws, zeilen_nr, BETRAG_SPALTE, gruppe["summe"], fett=True)
        ws.cell(row=zeilen_nr, column=17, value="x")
        for spalte in range(1, len(SPALTEN) + 1):
            ws.cell(row=zeilen_nr, column=spalte).fill = SUMME_FUELLUNG
        gesamt += gruppe["summe"]
        zeilen_nr += 1

    ws.cell(row=zeilen_nr, column=9, value="Gesamtsumme").font = SUMME_SCHRIFT
    _betrag_zelle(ws, zeilen_nr, BETRAG_SPALTE, gesamt, fett=True)
    for spalte in range(1, len(SPALTEN) + 1):
        ws.cell(row=zeilen_nr, column=spalte).fill = SUMME_FUELLUNG

    # --- Zusammenfassung ----------------------------------------------------
    daten = _zusammenfassung(fall, vorgang, list(abschlaege))
    zeilen_nr += 3

    def block_zeile(beschriftung, wert, konto="", fett=False):
        nonlocal zeilen_nr
        if konto:
            ws.cell(row=zeilen_nr, column=8, value=konto).font = SUMME_SCHRIFT if fett else Font()
        zelle = ws.cell(row=zeilen_nr, column=9, value=beschriftung)
        if fett:
            zelle.font = SUMME_SCHRIFT
        if wert is not None:
            _betrag_zelle(ws, zeilen_nr, BETRAG_SPALTE, wert, fett=fett)
        for spalte in range(8, 11):
            ws.cell(row=zeilen_nr, column=spalte).fill = BLOCK_FUELLUNG
        zeilen_nr += 1

    block_zeile("Bezüge DiakonInnen", fall["bezuege"], konto="603100")
    block_zeile("Umlage Versorgungsfonds", fall["umlage"], konto="609800")
    for konto, betrag in sorted(fall["konten"].items()):
        if konto not in ("603100", "609800"):
            block_zeile(f"Konto {konto}", betrag, konto=konto)
    if daten["kv"] is not None:
        block_zeile("Zuschuß KV Beihilfeber.", daten["kv"])
    block_zeile(
        "Beihilfeversicherung" if daten["beihilfe"] is not None else "Beihilfe (keine Kalkulation vorhanden!)",
        daten["beihilfe"],
    )
    block_zeile(f"PK {import_lauf.jahr} Gesamt:", daten["pk_gesamt"], fett=True)
    zeilen_nr += 1

    block_zeile("Bezüge DiakonInnen", fall["bezuege"], konto="10159")
    block_zeile("Umlage Versorgungsfonds", fall["umlage"], konto="10167")
    block_zeile("Beihilfeversicherungspauschale/KV-Zuschuss", daten["pos_beihilfe"], konto="10171")
    zeilen_nr += 1

    einzel = ", ".join(str(a.betrag) for a in abschlaege if a.betrag is not None)
    block_zeile(
        f"Abschlagszahlungen ({daten['abschlag_anzahl']} Stück{': ' + einzel if einzel else ''})",
        daten["abschlag_summe"],
    )
    block_zeile("Restbetrag:", daten["rest"], fett=True)

    return ws


def _vorgang_und_abschlaege(fall, jahr):
    einsatz = fall["einsatz"]
    if einsatz is None:
        return None, []
    vorgang = Fakturierungsvorgang.objects.filter(einsatz=einsatz, jahr=jahr).first()
    abschlaege = list(
        Quartalsabrechnung.objects.filter(einsatz=einsatz, jahr=jahr, quartal__lte=3)
        .order_by("quartal")
    )
    return vorgang, abschlaege


def einzel_mappe(import_lauf, fall):
    wb = Workbook()
    wb.remove(wb.active)
    vorgang, abschlaege = _vorgang_und_abschlaege(fall, import_lauf.jahr)
    fall_blatt(wb, import_lauf, fall, vorgang, abschlaege)
    puffer = io.BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    return puffer


STATUS_LABEL = {
    "offen": "Offen",
    "teilweise": "Teilweise",
    "ausgeglichen": "Ausgeglichen",
}


def cockpit_mappe(jahr, zeilen, summen, stufen):
    """Jahres-Cockpit als Excel: Pipeline-Status + Beträge je Abrechnungsfall."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Cockpit {jahr}"

    ws.cell(row=1, column=1, value=f"Jahres-Cockpit {jahr} – Abrechnungsstatus").font = TITEL_SCHRIFT

    stufen_labels = [label for _, label in stufen]
    kopf = ["Mitarbeiter", "PersNr", "Stelle", "Innenauftrag", "Debitor"] + stufen_labels + [
        "Status", "Hochrechnung", "Abschläge", "Rest", "Bezahlt", "Offen",
    ]
    kopf_zeile = 3
    for spalte, beschriftung in enumerate(kopf, start=1):
        zelle = ws.cell(row=kopf_zeile, column=spalte, value=beschriftung)
        zelle.font = KOPF_SCHRIFT
        zelle.fill = KOPF_FUELLUNG
    breiten = [22, 8, 22, 15, 24] + [7] * len(stufen_labels) + [13, 13, 13, 13, 13, 13]
    for spalte, breite in enumerate(breiten, start=1):
        ws.column_dimensions[get_column_letter(spalte)].width = breite
    ws.freeze_panes = "A4"

    betrag_start = 6 + len(stufen_labels) + 1  # erste Betragsspalte (Hochrechnung)
    zeilen_nr = kopf_zeile + 1
    for eintrag in zeilen:
        einsatz = eintrag["einsatz"]
        werte = [
            f"{einsatz.mitarbeiter.nachname}, {einsatz.mitarbeiter.vorname}",
            einsatz.mitarbeiter.personalnummer,
            einsatz.stelle.name,
            eintrag["innenauftrag"].nummer if eintrag["innenauftrag"] else "",
            str(eintrag["debitor"]) if eintrag["debitor"] else "",
        ]
        werte += ["✓" if eintrag["stufen"][key] else "–" for key, _ in stufen]
        werte.append(STATUS_LABEL.get(eintrag["status"], ""))
        for spalte, wert in enumerate(werte, start=1):
            zelle = ws.cell(row=zeilen_nr, column=spalte, value=wert)
            if 6 <= spalte <= 5 + len(stufen_labels) + 1:
                zelle.alignment = Alignment(horizontal="center")
        for offset, feld in enumerate(("hochrechnung", "abschlaege", "spitze_rest", "bezahlt", "offen")):
            _betrag_zelle(ws, zeilen_nr, betrag_start + offset, eintrag[feld])
        if zeilen_nr % 2 == 0:
            for spalte in range(1, len(kopf) + 1):
                ws.cell(row=zeilen_nr, column=spalte).fill = BLOCK_FUELLUNG
        zeilen_nr += 1

    ws.cell(row=zeilen_nr, column=1, value=f"Summe ({summen['faelle']} Fälle)").font = SUMME_SCHRIFT
    for offset, feld in enumerate(("hochrechnung", "abschlaege", "spitze_rest", "bezahlt", "offen")):
        _betrag_zelle(ws, zeilen_nr, betrag_start + offset, summen[feld], fett=True)
    for spalte in range(1, len(kopf) + 1):
        ws.cell(row=zeilen_nr, column=spalte).fill = SUMME_FUELLUNG

    puffer = io.BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    return puffer


def debitor_mappe(import_lauf, debitor, faelle_liste):
    """Sammelmappe: Deckblatt mit Übersicht + ein Blatt je Fall."""
    wb = Workbook()
    deckblatt = wb.active
    deckblatt.title = "Übersicht"

    deckblatt.cell(row=1, column=1, value=f"PK-Abrechnung {import_lauf.jahr} – {debitor.name}").font = TITEL_SCHRIFT
    zeile = 2
    for text in filter(None, [debitor.name2, debitor.anschriftperson, debitor.strasse,
                              f"{debitor.plz} {debitor.ort}".strip(),
                              f"Debitor: {debitor.sap_nummer}" if debitor.sap_nummer else ""]):
        deckblatt.cell(row=zeile, column=1, value=text)
        zeile += 1
    zeile += 1

    kopf = ["Fall", "PersNr", "Innenauftrag", "Bezüge (10159)", "Umlage (10167)",
            "Beihilfe/KV (10171)", "PK gesamt", "Abschläge", "Restbetrag"]
    for spalte, beschriftung in enumerate(kopf, start=1):
        zelle = deckblatt.cell(row=zeile, column=spalte, value=beschriftung)
        zelle.font = KOPF_SCHRIFT
        zelle.fill = KOPF_FUELLUNG
    breiten = [24, 8, 15, 14, 14, 16, 14, 14, 14]
    for spalte, breite in enumerate(breiten, start=1):
        deckblatt.column_dimensions[get_column_letter(spalte)].width = breite
    zeile += 1

    gesamt = {"bezuege": Decimal("0"), "umlage": Decimal("0"), "beihilfe": Decimal("0"),
              "pk": Decimal("0"), "abschlaege": Decimal("0"), "rest": Decimal("0")}

    for fall in faelle_liste:
        vorgang, abschlaege = _vorgang_und_abschlaege(fall, import_lauf.jahr)
        daten = _zusammenfassung(fall, vorgang, abschlaege)
        werte = [
            f"{fall['nachname']}, {fall['vorname']}", fall["personalnummer"], fall["auftrag_nummer"],
            fall["bezuege"], fall["umlage"], daten["pos_beihilfe"],
            daten["pk_gesamt"], daten["abschlag_summe"], daten["rest"],
        ]
        for spalte, wert in enumerate(werte, start=1):
            zelle = deckblatt.cell(row=zeile, column=spalte, value=wert)
            if spalte >= 4:
                zelle.number_format = BETRAG_FORMAT
        if zeile % 2 == 0:
            for spalte in range(1, len(kopf) + 1):
                deckblatt.cell(row=zeile, column=spalte).fill = BLOCK_FUELLUNG
        zeile += 1
        gesamt["bezuege"] += fall["bezuege"]
        gesamt["umlage"] += fall["umlage"]
        gesamt["beihilfe"] += daten["pos_beihilfe"]
        gesamt["pk"] += daten["pk_gesamt"]
        gesamt["abschlaege"] += daten["abschlag_summe"]
        gesamt["rest"] += daten["rest"]

        fall_blatt(wb, import_lauf, fall, vorgang, abschlaege)

    deckblatt.cell(row=zeile, column=1, value="Summe").font = SUMME_SCHRIFT
    for spalte, wert in enumerate(
        [gesamt["bezuege"], gesamt["umlage"], gesamt["beihilfe"], gesamt["pk"], gesamt["abschlaege"], gesamt["rest"]],
        start=4,
    ):
        _betrag_zelle(deckblatt, zeile, spalte, wert, fett=True)
    for spalte in range(1, len(kopf) + 1):
        deckblatt.cell(row=zeile, column=spalte).fill = SUMME_FUELLUNG

    puffer = io.BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    return puffer
