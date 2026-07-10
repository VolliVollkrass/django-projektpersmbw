"""Excel-Mappen lesen (Import) und schreiben (Export) im Vorlagenformat.

Aufbau eines Blatts: Zeile 1 = PFLICHT/optional-Marker, Zeile 2 = Überschriften,
danach Daten. Mappen ohne Markerzeile (Überschriften in Zeile 1) werden ebenso
akzeptiert – zugeordnet wird immer über die Spaltenüberschrift, nicht die
Position.
"""

import io
from datetime import date, datetime

import openpyxl
import tablib
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .registry import SHEETS

MARKER_WERTE = {"pflicht", "optional"}


def _zelle_normalisieren(wert):
    """Excel-Zellwerte in importfreundliche Python-Werte wandeln."""
    if wert is None:
        return ""
    if isinstance(wert, str):
        return wert.strip()
    if isinstance(wert, float) and wert.is_integer():
        return int(wert)
    if isinstance(wert, datetime):
        return wert.date()
    return wert


def _ist_markerzeile(zeile):
    werte = [str(z).strip().lower() for z in zeile if z not in (None, "")]
    return bool(werte) and all(w in MARKER_WERTE for w in werte)


def mappe_lesen(datei):
    """Liest eine hochgeladene Mappe.

    Rückgabe: Liste von (SheetConfig, tablib.Dataset, kopf_offset) für alle
    bekannten Blätter mit Datenzeilen, in Import-Reihenfolge. kopf_offset ist
    die Excel-Zeilennummer der Überschriftenzeile (für Fehlermeldungen).
    """
    wb = openpyxl.load_workbook(datei, data_only=True, read_only=True)
    ergebnis = []
    for config in SHEETS:
        if config.sheet_name not in wb.sheetnames:
            continue
        ws = wb[config.sheet_name]
        zeilen = [[_zelle_normalisieren(z) for z in zeile] for zeile in ws.iter_rows(values_only=True)]
        # Zeilen am Ende ohne Inhalt entfernen
        while zeilen and all(z == "" for z in zeilen[-1]):
            zeilen.pop()
        if not zeilen:
            continue

        kopf_index = 1 if _ist_markerzeile(zeilen[0]) else 0
        if len(zeilen) <= kopf_index:
            continue
        kopf = [str(z).strip() for z in zeilen[kopf_index]]
        # Leere Spalten am Ende abschneiden
        while kopf and kopf[-1] == "":
            kopf.pop()
        if not kopf:
            continue

        daten = tablib.Dataset()
        daten.headers = kopf
        for zeile in zeilen[kopf_index + 1:]:
            werte = list(zeile[: len(kopf)])
            werte += [""] * (len(kopf) - len(werte))
            if all(w == "" for w in werte):
                continue
            daten.append(werte)

        if len(daten) > 0:
            ergebnis.append((config, daten, kopf_index + 1))

    wb.close()
    return ergebnis


# ELKB-CI-Farben (wie das App-Theme in assets/css/input.css)
ELKB_VIOLETT = "5B2281"       # primary
ELKB_BLAUGRAU = "839ABA"      # secondary
ELKB_VIOLETT_MITTEL = "C8B8DB"  # base-200
ELKB_VIOLETT_HELL = "ECE7F3"    # base-100

MARKER_FUELLUNG = PatternFill("solid", fgColor=ELKB_BLAUGRAU)
MARKER_SCHRIFT = Font(bold=True, color="FFFFFF")
KOPF_FUELLUNG = PatternFill("solid", fgColor=ELKB_VIOLETT)
KOPF_SCHRIFT = Font(bold=True, color="FFFFFF")
SCHLUESSEL_FUELLUNG = PatternFill("solid", fgColor=ELKB_VIOLETT_MITTEL)
SCHLUESSEL_SCHRIFT = Font(bold=True, color=ELKB_VIOLETT)
STREIFEN_FUELLUNG = PatternFill("solid", fgColor=ELKB_VIOLETT_HELL)


def mappe_schreiben(datensaetze):
    """Erzeugt die Export-Mappe.

    datensaetze: Liste von (SheetConfig, tablib.Dataset) – das Dataset kommt aus
    resource.export() und trägt die Vorlagen-Überschriften. Die Datei ist
    direkt wieder importierbar (Markerzeile + Überschriften wie die Vorlage).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for config, daten in datensaetze:
        ws = wb.create_sheet(title=config.sheet_name)
        kopf = list(daten.headers or [])

        marker = [
            "PFLICHT" if spalte in config.pflicht_spalten else "optional"
            for spalte in kopf
        ]
        ws.append(marker)
        ws.append(kopf)
        for zeile in daten:
            ws.append(["" if w is None else w for w in zeile])

        for spalte_nr, spalte in enumerate(kopf, start=1):
            marker_zelle = ws.cell(row=1, column=spalte_nr)
            marker_zelle.font = MARKER_SCHRIFT
            marker_zelle.fill = MARKER_FUELLUNG
            kopf_zelle = ws.cell(row=2, column=spalte_nr)
            if spalte in config.pflicht_spalten:
                kopf_zelle.font = SCHLUESSEL_SCHRIFT
                kopf_zelle.fill = SCHLUESSEL_FUELLUNG
            else:
                kopf_zelle.font = KOPF_SCHRIFT
                kopf_zelle.fill = KOPF_FUELLUNG
            breite = max([len(str(spalte))] + [len(str(z)) for z in daten[spalte] if z is not None][:200] or [10])
            ws.column_dimensions[get_column_letter(spalte_nr)].width = min(max(breite + 2, 12), 45)

        # Dezente Zeilenstreifen in ELKB-Hellviolett
        for zeilen_nr in range(4, ws.max_row + 1, 2):
            for spalte_nr in range(1, len(kopf) + 1):
                ws.cell(row=zeilen_nr, column=spalte_nr).fill = STREIFEN_FUELLUNG

        ws.freeze_panes = "A3"

    puffer = io.BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    return puffer
