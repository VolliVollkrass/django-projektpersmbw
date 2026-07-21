import io
from datetime import date
from decimal import Decimal

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase

from ansprechpartner.models import Ansprechpartner
from einsatz.models import Einsatz
from personal.models import Mitarbeiter
from traeger.models import Einrichtung, Stelle, Traeger

from .registry import SHEETS
from .services import import_ausfuehren
from .xlsx import mappe_lesen, mappe_schreiben


def mappe_bauen(blaetter, mit_marker=True):
    """Baut eine Test-Mappe: blaetter = {Blattname: (kopf, [zeilen])}."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, (kopf, zeilen) in blaetter.items():
        ws = wb.create_sheet(title=name)
        if mit_marker:
            ws.append(["PFLICHT"] * len(kopf))
        ws.append(kopf)
        for zeile in zeilen:
            ws.append(zeile)
    puffer = io.BytesIO()
    wb.save(puffer)
    puffer.seek(0)
    return puffer


VOLLE_MAPPE = {
    "Ansprechpartner": (
        ["Vorname", "Nachname", "Geschlecht", "E-Mail", "Telefon", "Notiz", "Aktiv"],
        [["Maria", "Muster", "weiblich", "m.muster@traeger.de", "0911 123456", "Vorstand", "Ja"]],
    ),
    "Träger": (
        ["Träger-Id", "Name", "Art", "Haupt-Ansprechpartner (E-Mail)", "Straße", "PLZ", "Ort", "Aktiv", "Bemerkung"],
        [["TR-001", "Diakonisches Werk Musterstadt", "Diakonischer Träger", "m.muster@traeger.de", "Hauptstr. 1", "90402", "Nürnberg", "Ja", ""]],
    ),
    "Einrichtungen": (
        ["Einrichtungs-Id", "Name", "Träger-Id", "Art", "Straße", "PLZ", "Ort", "Ansprechpartner (E-Mail)", "Versorgungsumlage", "Aktiv", "Bemerkung"],
        [
            ["16393", "Kita Sonnenschein", "TR-001", "Diakonische Einrichtung", "Gartenweg 5", "90402", "Nürnberg", "", "40 %", "Ja", ""],
            ["16393", "Jugendhaus Arche", "TR-001", "Diakonische Einrichtung", "", "", "", "", "keine", "Ja", ""],
        ],
    ),
    "Stellen": (
        ["Stellen-Id", "Name", "Einrichtungs-Id", "Einrichtung (Name)", "Position", "Straße", "PLZ", "Ort", "Ansprechpartner (E-Mail)", "Aktiv", "Bemerkung"],
        [["ST-001", "Gemeindediakon", "16393", "Kita Sonnenschein", "Jugendarbeit", "", "", "", "", "Ja", ""]],
    ),
    "Mitarbeiter": (
        ["Personalnummer", "Geschlecht", "Vorname", "Nachname", "Geburtsdatum", "E-Mail", "Telefon privat", "Telefon dienstlich", "Straße", "PLZ", "Ort", "Anstellung", "Dienststand", "Besoldungsgruppe", "Stufe", "Familienstand", "Kinder", "Kindergeldberechtigt", "Voller OFZ-Anspruch", "Aktiv", "Bemerkung"],
        [[10001, "männlich", "Thomas", "Beispiel", "1985-06-15", "t.beispiel@example.de", "", "", "", 90403, "Nürnberg", "öffentlich-rechtlich", "Diakon", "A10", 4, "Verheiratet", 2, "Ja", "Nein", "Ja", ""]],
    ),
    "Einsätze": (
        ["Stellen-Id", "Stelle (Name)", "Personalnummer", "Beginn", "Ende", "Umfang", "Abrechnung über LKA", "Aktiv", "Bemerkung"],
        [["ST-001", "", 10001, "2026-01-01", "", "0,5", "Ja", "Ja", ""]],
    ),
}


class ImportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user("sachbearbeitung", password="x" * 12)

    def test_kompletter_import_mit_marker_und_klartext(self):
        ergebnisse, gespeichert = import_ausfuehren(
            mappe_bauen(VOLLE_MAPPE), self.user, commit=True
        )
        self.assertTrue(gespeichert, [e.fehler for e in ergebnisse])
        self.assertEqual(Traeger.objects.count(), 1)
        self.assertEqual(Einrichtung.objects.count(), 2)

        traeger = Traeger.objects.get(traeger_id="TR-001")
        self.assertEqual(traeger.art, "dw")  # Label wurde in Code übersetzt
        self.assertEqual(traeger.haupt_ansprechpartner.email, "m.muster@traeger.de")

        kita = Einrichtung.objects.get(einrichtungs_id="16393", name="Kita Sonnenschein")
        self.assertEqual(kita.versorgungsumlage, Decimal("0.40"))
        self.assertEqual(kita.plz, "90402")

        stelle = Stelle.objects.get(stellen_id="ST-001")
        self.assertEqual(stelle.einrichtung, kita)  # Sammel-Id über Name aufgelöst

        mitarbeiter = Mitarbeiter.objects.get(personalnummer="10001")
        self.assertEqual(mitarbeiter.geburtsdatum, date(1985, 6, 15))
        self.assertEqual(mitarbeiter.plz, "90403")
        self.assertTrue(mitarbeiter.kindergeldberechtigt)
        self.assertEqual(mitarbeiter.angelegt_von, self.user)

        einsatz = Einsatz.objects.get()
        self.assertEqual(einsatz.stelle, stelle)
        self.assertEqual(einsatz.umfang, Decimal("0.50"))
        self.assertTrue(einsatz.abrechnung)

    def test_reimport_ist_idempotent(self):
        import_ausfuehren(mappe_bauen(VOLLE_MAPPE), self.user, commit=True)
        ergebnisse, gespeichert = import_ausfuehren(
            mappe_bauen(VOLLE_MAPPE), self.user, commit=True
        )
        self.assertTrue(gespeichert)
        for e in ergebnisse:
            self.assertEqual(e.neu, 0, f"{e.sheet_name}: {e.fehler}")
            self.assertEqual(e.aktualisiert, 0, f"{e.sheet_name} wurde geändert")
        self.assertEqual(Einrichtung.objects.count(), 2)
        self.assertEqual(Einsatz.objects.count(), 1)

    def test_aenderung_wird_als_update_erkannt(self):
        import_ausfuehren(mappe_bauen(VOLLE_MAPPE), self.user, commit=True)
        geaendert = dict(VOLLE_MAPPE)
        kopf, zeilen = geaendert["Träger"]
        neue_zeile = list(zeilen[0])
        neue_zeile[4] = "Neue Straße 99"
        geaendert["Träger"] = (kopf, [neue_zeile])

        ergebnisse, gespeichert = import_ausfuehren(mappe_bauen(geaendert), self.user, commit=True)
        self.assertTrue(gespeichert)
        blatt = {e.sheet_name: e for e in ergebnisse}["Träger"]
        self.assertEqual((blatt.neu, blatt.aktualisiert), (0, 1))
        self.assertEqual(Traeger.objects.get().strasse, "Neue Straße 99")
        self.assertEqual(Traeger.objects.count(), 1)

    def test_sammel_id_ohne_namen_ist_fehler(self):
        mappe = dict(VOLLE_MAPPE)
        kopf, zeilen = mappe["Stellen"]
        zeile = list(zeilen[0])
        zeile[3] = ""  # 'Einrichtung (Name)' leeren → Sammel-Id 16393 mehrdeutig
        mappe["Stellen"] = (kopf, [zeile])

        ergebnisse, gespeichert = import_ausfuehren(mappe_bauen(mappe), self.user, commit=True)
        self.assertFalse(gespeichert)
        blatt = {e.sheet_name: e for e in ergebnisse}["Stellen"]
        self.assertTrue(blatt.hat_fehler)
        self.assertIn("Sammel-Id", blatt.fehler[0][1])
        # Alles-oder-nichts: auch die fehlerfreien Blätter wurden nicht gespeichert
        self.assertEqual(Traeger.objects.count(), 0)

    def test_fehler_verhindert_gesamten_import(self):
        mappe = dict(VOLLE_MAPPE)
        kopf, zeilen = mappe["Mitarbeiter"]
        zeile = list(zeilen[0])
        zeile[4] = "kein-datum"
        mappe["Mitarbeiter"] = (kopf, [zeile])

        ergebnisse, gespeichert = import_ausfuehren(mappe_bauen(mappe), self.user, commit=True)
        self.assertFalse(gespeichert)
        self.assertEqual(Mitarbeiter.objects.count(), 0)
        self.assertEqual(Traeger.objects.count(), 0)

    def test_mappe_ohne_markerzeile_wird_akzeptiert(self):
        nur_traeger = {"Träger": VOLLE_MAPPE["Träger"], "Ansprechpartner": VOLLE_MAPPE["Ansprechpartner"]}
        ergebnisse, gespeichert = import_ausfuehren(
            mappe_bauen(nur_traeger, mit_marker=False), self.user, commit=True
        )
        self.assertTrue(gespeichert, [e.fehler for e in ergebnisse])
        self.assertEqual(Traeger.objects.count(), 1)

    def test_vorschau_schreibt_nichts(self):
        ergebnisse, gespeichert = import_ausfuehren(
            mappe_bauen(VOLLE_MAPPE), self.user, commit=False
        )
        self.assertFalse(gespeichert)
        self.assertFalse(any(e.hat_fehler for e in ergebnisse), [e.fehler for e in ergebnisse])
        blatt = {e.sheet_name: e for e in ergebnisse}["Träger"]
        self.assertEqual(blatt.neu, 1)
        self.assertEqual(Traeger.objects.count(), 0)  # Rollback

    def test_export_roundtrip_ist_unveraendert(self):
        import_ausfuehren(mappe_bauen(VOLLE_MAPPE), self.user, commit=True)

        datensaetze = [(config, config.resource_class().export()) for config in SHEETS]
        export_datei = mappe_schreiben(datensaetze)

        ergebnisse, gespeichert = import_ausfuehren(export_datei, self.user, commit=True)
        self.assertTrue(gespeichert, [e.fehler for e in ergebnisse])
        for e in ergebnisse:
            self.assertEqual(e.neu, 0, f"{e.sheet_name}: {e.fehler}")
            self.assertEqual(e.aktualisiert, 0, f"{e.sheet_name} unterscheidet sich vom Export")

    def test_vorlagen_datei_ist_lesbar(self):
        from .views import VORLAGE_PFAD

        with open(VORLAGE_PFAD, "rb") as f:
            blaetter = mappe_lesen(io.BytesIO(f.read()))
        # Beispielzeile (Zeile 3) jedes Blatts wird als Datenzeile erkannt
        namen = [config.sheet_name for config, _, _ in blaetter]
        self.assertIn("Träger", namen)
        self.assertIn("Einsätze", namen)


class MbwStammdatenRoundtripTests(TestCase):
    """Debitoren + Innenaufträge nehmen roundtrip-fähig am Export/Import teil."""

    @classmethod
    def setUpTestData(cls):
        from mbw.models import Debitor, Innenauftrag

        cls.user = User.objects.create_user("sachbearbeitung", password="x" * 12)
        cls.traeger = Traeger.objects.create(traeger_id="TR-9", name="Diakonie X", art="dw")
        einrichtung = Einrichtung.objects.create(
            einrichtungs_id="E-9", name="Haus Nord", traeger=cls.traeger
        )
        cls.stelle = Stelle.objects.create(
            stellen_id="S-9", name="Gemeindediakon", einrichtung=einrichtung
        )
        cls.mitarbeiter = Mitarbeiter.objects.create(
            personalnummer="20001", vorname="Anna", nachname="Test",
            geburtsdatum=date(1980, 1, 1),
        )
        cls.einsatz = Einsatz.objects.create(
            stelle=cls.stelle, mitarbeiter=cls.mitarbeiter, beginn=date(2026, 1, 1)
        )
        cls.debitor = Debitor.objects.create(
            name="Rechnungsanschrift X", sap_nummer="1900099999", ort="Nürnberg",
            traeger=cls.traeger, versandweg=Debitor.Versandweg.EMAIL,
        )
        cls.innenauftrag = Innenauftrag.objects.create(
            nummer="F099999-0001", kostenstelle="3-0312-033", bezeichnung="Testauftrag",
            debitor=cls.debitor, einsatz=cls.einsatz, andock_art="stelle",
        )

    def _export_mbw(self):
        configs = [c for c in SHEETS if c.sheet_name in ("Debitoren", "Innenaufträge")]
        datensaetze = [(c, c.resource_class().export()) for c in configs]
        return mappe_schreiben(datensaetze)

    def test_roundtrip_ist_idempotent(self):
        ergebnisse, gespeichert = import_ausfuehren(self._export_mbw(), self.user, commit=True)
        self.assertTrue(gespeichert, [e.fehler for e in ergebnisse])
        blaetter = {e.sheet_name: e for e in ergebnisse}
        for name in ("Debitoren", "Innenaufträge"):
            self.assertEqual(blaetter[name].neu, 0, f"{name}: {blaetter[name].fehler}")
            self.assertEqual(blaetter[name].aktualisiert, 0, f"{name} unterscheidet sich vom Export")
        # Einsatz-Bezug am Innenauftrag bleibt beim Import unangetastet
        self.innenauftrag.refresh_from_db()
        self.assertEqual(self.innenauftrag.einsatz, self.einsatz)
        self.assertEqual(self.innenauftrag.andock_art, "stelle")

    def test_export_enthaelt_einsatz_infospalten(self):
        from .xlsx import mappe_lesen

        blaetter = mappe_lesen(self._export_mbw())
        ia_config, daten, _ = next(b for b in blaetter if b[0].sheet_name == "Innenaufträge")
        zeile = daten.dict[0]
        self.assertEqual(zeile["Einsatz (Personalnr.)"], "20001")
        self.assertEqual(zeile["Einsatz (Stellen-Id)"], "S-9")
        self.assertEqual(zeile["Debitor-Id"], self.debitor.pk)

    def test_debitor_aenderung_wird_als_update_erkannt(self):
        from mbw.models import Debitor

        import_ausfuehren(self._export_mbw(), self.user, commit=True)
        # Ort im Export ändern und erneut importieren
        wb = openpyxl.load_workbook(self._export_mbw())
        ws = wb["Debitoren"]
        kopf = [c.value for c in ws[2]]
        ort_spalte = kopf.index("Ort") + 1
        ws.cell(row=3, column=ort_spalte, value="München")
        puffer = io.BytesIO(); wb.save(puffer); puffer.seek(0)

        ergebnisse, gespeichert = import_ausfuehren(puffer, self.user, commit=True)
        self.assertTrue(gespeichert, [e.fehler for e in ergebnisse])
        blatt = {e.sheet_name: e for e in ergebnisse}["Debitoren"]
        self.assertEqual((blatt.neu, blatt.aktualisiert), (0, 1))
        self.debitor.refresh_from_db()
        self.assertEqual(self.debitor.ort, "München")
        self.assertEqual(Debitor.objects.count(), 1)
